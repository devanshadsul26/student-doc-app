"""
services/benchmark_service.py

Generates real test files of various sizes, uploads them to BOTH Cloud SQL
and GCS, measures actual timings, and returns results ready for Excel export.
"""

import io
import os
import random
import string
from db.queries import create_student, insert_metadata, insert_blob_timed
from db.queries import fetch_blob_timed
from storage.gcs import upload_file_timed, download_file_timed
from utils.cost_calculator import estimate_cost

# Benchmark student used for all test uploads
BENCHMARK_STUDENT_ID = "BENCHMARK_TEST"
BENCHMARK_STUDENT_NAME = "Benchmark Runner"

# File sizes to test: (label, size_in_bytes)
BENCHMARK_SIZES = [
    ("1 KB",    1 * 1024),
    ("10 KB",   10 * 1024),
    ("50 KB",   50 * 1024),
    ("100 KB",  100 * 1024),
    ("250 KB",  250 * 1024),
    ("500 KB",  500 * 1024),
    ("1 MB",    1 * 1024 * 1024),
    ("2 MB",    2 * 1024 * 1024),
    ("5 MB",    5 * 1024 * 1024),
]


def _generate_file_bytes(size_bytes: int) -> bytes:
    """Generate random bytes to simulate a real file of the given size."""
    # Use random printable ASCII to make it realistic (like a text/PDF body)
    chunk = (string.ascii_letters + string.digits + " \n").encode()
    repeats = (size_bytes // len(chunk)) + 1
    return (chunk * repeats)[:size_bytes]


def run_benchmark(runs_per_size: int = 3, progress_callback=None) -> list[dict]:
    """
    For each file size in BENCHMARK_SIZES, upload `runs_per_size` times to
    both Cloud SQL and GCS, measure real upload + download times, and return
    a list of result dicts.

    progress_callback(current, total, label) — optional UI progress hook.
    """
    create_student(BENCHMARK_STUDENT_ID, BENCHMARK_STUDENT_NAME)

    results = []
    total_ops = len(BENCHMARK_SIZES) * runs_per_size
    op = 0

    for size_label, size_bytes in BENCHMARK_SIZES:
        for run in range(1, runs_per_size + 1):
            op += 1
            filename = f"bench_{size_label.replace(' ', '')}_{run}.bin"
            gcs_path = f"benchmark/{BENCHMARK_STUDENT_ID}/{filename}"

            if progress_callback:
                progress_callback(op, total_ops, f"{size_label} — run {run}/{runs_per_size}")

            file_bytes = _generate_file_bytes(size_bytes)

            # ── Upload to Cloud SQL ──
            sql_upload_ms = insert_blob_timed(
                BENCHMARK_STUDENT_ID, "Benchmark", filename, file_bytes
            )

            # ── Upload to GCS ──
            _, gcs_upload_ms = upload_file_timed(io.BytesIO(file_bytes), gcs_path)

            # Save GCS metadata
            insert_metadata(
                BENCHMARK_STUDENT_ID, "Benchmark", filename,
                gcs_path, size_bytes
            )

            # ── Download from Cloud SQL ──
            _, sql_download_ms = fetch_blob_timed(BENCHMARK_STUDENT_ID, filename)

            # ── Download from GCS ──
            _, gcs_download_ms = download_file_timed(gcs_path)

            # ── Cost ──
            cost = estimate_cost(size_bytes)

            results.append({
                "size_label": size_label,
                "size_bytes": size_bytes,
                "size_kb": round(size_bytes / 1024, 2),
                "run": run,
                "sql_upload_ms": sql_upload_ms,
                "gcs_upload_ms": gcs_upload_ms,
                "sql_download_ms": sql_download_ms,
                "gcs_download_ms": gcs_download_ms,
                "sql_cost_usd": cost["sql_monthly_usd"],
                "gcs_cost_usd": cost["gcs_monthly_usd"],
                "faster_upload": "SQL" if sql_upload_ms < gcs_upload_ms else "GCS",
                "faster_download": "SQL" if sql_download_ms < gcs_download_ms else "GCS",
            })

    return results


def results_to_excel(results: list[dict]) -> bytes:
    """Convert benchmark results to an Excel file and return as bytes."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()

    # ── Sheet 1: Raw Results ──
    ws_raw = wb.active
    ws_raw.title = "Raw Results"

    headers = [
        "Size", "Size (KB)", "Run",
        "SQL Upload (ms)", "GCS Upload (ms)",
        "SQL Download (ms)", "GCS Download (ms)",
        "SQL Cost/mo ($)", "GCS Cost/mo ($)",
        "Faster Upload", "Faster Download"
    ]

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws_raw.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, r in enumerate(results, 2):
        values = [
            r["size_label"], r["size_kb"], r["run"],
            r["sql_upload_ms"], r["gcs_upload_ms"],
            r["sql_download_ms"], r["gcs_download_ms"],
            r["sql_cost_usd"], r["gcs_cost_usd"],
            r["faster_upload"], r["faster_download"],
        ]
        for col, val in enumerate(values, 1):
            ws_raw.cell(row=row_idx, column=col, value=val)

        # Colour faster upload column
        fu_cell = ws_raw.cell(row=row_idx, column=10)
        fd_cell = ws_raw.cell(row=row_idx, column=11)
        for cell in [fu_cell, fd_cell]:
            if cell.value == "GCS":
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                cell.font = Font(color="276221")
            else:
                cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                cell.font = Font(color="9C0006")

    for col in ws_raw.columns:
        ws_raw.column_dimensions[col[0].column_letter].width = max(
            len(str(c.value or "")) for c in col
        ) + 4

    # ── Sheet 2: Averages per size ──
    ws_avg = wb.create_sheet("Averages by Size")

    avg_headers = [
        "Size", "Size (KB)",
        "Avg SQL Upload (ms)", "Avg GCS Upload (ms)",
        "Avg SQL Download (ms)", "Avg GCS Download (ms)",
        "SQL Cost/mo ($)", "GCS Cost/mo ($)",
        "Upload Winner", "Download Winner"
    ]
    for col, h in enumerate(avg_headers, 1):
        cell = ws_avg.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Group by size_label
    from collections import defaultdict
    groups = defaultdict(list)
    for r in results:
        groups[r["size_label"]].append(r)

    size_order = [s[0] for s in BENCHMARK_SIZES]
    for row_idx, size_label in enumerate(size_order, 2):
        group = groups.get(size_label, [])
        if not group:
            continue

        def avg(key):
            return round(sum(r[key] for r in group) / len(group), 2)

        avg_sql_up = avg("sql_upload_ms")
        avg_gcs_up = avg("gcs_upload_ms")
        avg_sql_dl = avg("sql_download_ms")
        avg_gcs_dl = avg("gcs_download_ms")

        values = [
            size_label, group[0]["size_kb"],
            avg_sql_up, avg_gcs_up,
            avg_sql_dl, avg_gcs_dl,
            group[0]["sql_cost_usd"], group[0]["gcs_cost_usd"],
            "SQL" if avg_sql_up < avg_gcs_up else "GCS",
            "SQL" if avg_sql_dl < avg_gcs_dl else "GCS",
        ]
        for col, val in enumerate(values, 1):
            ws_avg.cell(row=row_idx, column=col, value=val)

    for col in ws_avg.columns:
        ws_avg.column_dimensions[col[0].column_letter].width = max(
            len(str(c.value or "")) for c in col
        ) + 4

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
