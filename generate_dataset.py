"""
generate_dataset.py
Run this once to generate a sample_dataset.xlsx with 50 simulated upload records.
Usage: python generate_dataset.py
"""

import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta

random.seed(42)

STUDENT_IDS = [f"S{str(i).zfill(3)}" for i in range(1, 11)]
STUDENT_NAMES = [
    "Alice Johnson", "Bob Smith", "Carol White", "David Brown", "Eva Martinez",
    "Frank Lee", "Grace Kim", "Henry Patel", "Iris Chen", "James Wilson"
]
DOC_TYPES = ["ID", "Transcript", "Certificate", "Other"]
FILENAMES = [
    "passport.pdf", "transcript_2024.pdf", "birth_certificate.pdf",
    "id_card.png", "degree_certificate.pdf", "marksheet.pdf",
    "recommendation_letter.pdf", "resume.pdf", "photo.jpg", "admission_form.pdf",
    "fee_receipt.pdf", "medical_certificate.pdf", "character_certificate.pdf",
    "migration_certificate.pdf", "noc_letter.pdf"
]

# Simulate realistic upload times (ms)
# SQL tends to be faster for small files but slower for large ones
def simulate_times(size_kb):
    base_sql = 80 + (size_kb * 0.3) + random.uniform(-20, 40)
    base_gcs = 120 + (size_kb * 0.15) + random.uniform(-30, 60)
    return round(base_sql, 2), round(base_gcs, 2)

GCS_PRICE = 0.023
SQL_PRICE = 0.17
BYTES_PER_GB = 1024 ** 3

rows = []
base_date = datetime(2026, 1, 1, 9, 0, 0)

for i in range(50):
    idx = i % 10
    student_id = STUDENT_IDS[idx]
    student_name = STUDENT_NAMES[idx]
    doc_type = random.choice(DOC_TYPES)
    filename = random.choice(FILENAMES)
    size_kb = round(random.uniform(50, 2000), 2)
    size_bytes = int(size_kb * 1024)
    sql_ms, gcs_ms = simulate_times(size_kb)
    sql_cost = round((size_bytes / BYTES_PER_GB) * SQL_PRICE, 8)
    gcs_cost = round((size_bytes / BYTES_PER_GB) * GCS_PRICE, 8)
    faster = "SQL" if sql_ms < gcs_ms else "GCS"
    upload_date = base_date + timedelta(hours=i * 3)

    rows.append({
        "upload_no": i + 1,
        "student_id": student_id,
        "student_name": student_name,
        "doc_type": doc_type,
        "filename": filename,
        "file_size_kb": size_kb,
        "sql_upload_ms": sql_ms,
        "gcs_upload_ms": gcs_ms,
        "sql_cost_usd": sql_cost,
        "gcs_cost_usd": gcs_cost,
        "faster": faster,
        "uploaded_at": upload_date.strftime("%Y-%m-%d %H:%M"),
    })

# ── Write to Excel ──
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Upload Benchmark"

headers = [
    "Upload #", "Student ID", "Student Name", "Doc Type", "Filename",
    "File Size (KB)", "SQL Upload (ms)", "GCS Upload (ms)",
    "SQL Cost/mo ($)", "GCS Cost/mo ($)", "Faster Service", "Uploaded At"
]

# Header row styling
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

# Data rows
for row_idx, row in enumerate(rows, 2):
    values = [
        row["upload_no"], row["student_id"], row["student_name"],
        row["doc_type"], row["filename"], row["file_size_kb"],
        row["sql_upload_ms"], row["gcs_upload_ms"],
        row["sql_cost_usd"], row["gcs_cost_usd"],
        row["faster"], row["uploaded_at"]
    ]
    for col, val in enumerate(values, 1):
        ws.cell(row=row_idx, column=col, value=val)

    # Highlight "Faster" column
    faster_cell = ws.cell(row=row_idx, column=11)
    if faster_cell.value == "GCS":
        faster_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        faster_cell.font = Font(color="276221")
    else:
        faster_cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        faster_cell.font = Font(color="9C0006")

# Auto column widths
for col in ws.columns:
    max_len = max(len(str(cell.value or "")) for cell in col)
    ws.column_dimensions[col[0].column_letter].width = max_len + 4

wb.save("sample_dataset.xlsx")
print("✅ sample_dataset.xlsx created with 50 upload records.")
